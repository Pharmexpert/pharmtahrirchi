"""
МИССИЯ Д — Фармакопея маълумотлар базасини СИФАТЛИ қуриш.

Архитектура:
1. Мундарижа (XLSX) → бўлим рақамлари (matn_no: 1.1., 2.1.1., ...)
2. MD файллардан matn_no бўйича бўлим матнини топиш (### header matching)
3. UZ + RU бўлимларни жуфтлаштириш (alignment)
4. Гапларга ажратиш → Хатбошилар
5. Терминлар чиқариш → Изоҳли луғат, Мунозарали, Қисқартмалар

Манбалар:
  XLSX: Qonun/Forms/ДФ мундарижаси жадвали 1 нашр, 1 жилд.xlsx
  MD UZ: Qonun/5/DF_uzb/1-жилд/1_қисм/, 2_қисм/, ...
  MD RU: Qonun/5/DF_rus/1-жилд/1_часть/, 2_часть/, ...
  MD EN: Qonun/5/EP_9.0/...

Натижа: Qonun/5/ папкасига XLSX файллар
"""
import os
import re
import sys
import logging
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')
log = logging.getLogger("mission_d")

BASE = Path(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
FORMS = BASE / "Qonun" / "Forms"
MD_DIR = BASE / "Qonun" / "5"
OUT_DIR = BASE / "Qonun" / "5"

# Styles
HF = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HBG = PatternFill(start_color='8B5E3C', end_color='8B5E3C', fill_type='solid')
CF = Font(name='Arial', size=10)
BD = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WR = Alignment(wrap_text=True, vertical='top')


def load_toc():
    """XLSX мундарижани юклаш."""
    p = FORMS / "ДФ мундарижаси жадвали 1 нашр, 1 жилд.xlsx"
    wb = openpyxl.load_workbook(str(p), data_only=True)
    ws = wb[wb.sheetnames[0]]
    entries = []
    for r in range(2, ws.max_row + 1):
        en = str(ws.cell(r, 3).value or '').strip()
        uz = str(ws.cell(r, 4).value or '').strip()
        ru = str(ws.cell(r, 5).value or '').strip()
        mn = str(ws.cell(r, 6).value or '').strip()
        jl = str(ws.cell(r, 7).value or '').strip()
        if en or uz or ru:
            entries.append({'en': en, 'uz': uz, 'ru': ru, 'matn_no': mn, 'jild': jl})
    wb.close()
    return entries


def load_all_md(lang_dir: str) -> dict:
    """MD файлларни юклаш ва бўлим рақамлари бўйича индекслаш.

    Returns: {section_number: full_text}
    Мисол: {'1.1.': 'Умумий қоидалар матни...', '2.1.1.': 'Каплемер матни...'}
    """
    sections = {}
    md_root = MD_DIR / lang_dir

    if not md_root.exists():
        return sections

    for md_file in sorted(md_root.rglob("*.md")):
        try:
            content = md_file.read_text(encoding='utf-8', errors='replace')
        except Exception:
            continue

        # Бўлим рақамлари бўйича split (### 2.1.1. Header)
        # Pattern: line starts with optional # and section number like 1.1. or 2.3.4.
        parts = re.split(r'\n(?=#{1,4}\s*\d+\.[\d.]*)', content)

        for part in parts:
            # Extract section number from first line
            m = re.match(r'#{1,4}\s*(\d+\.[\d.]*)\s*(.*)', part.strip())
            if m:
                sec_no = m.group(1).rstrip('.')
                # Normalize: 1.1 → 1.1., 2.1.1 → 2.1.1.
                if not sec_no.endswith('.'):
                    sec_no += '.'
                # Clean text: remove markdown headers, metadata
                text = part.strip()
                text = re.sub(r'^#{1,4}\s*', '', text, count=1)
                text = re.sub(r'^>.*$', '', text, flags=re.MULTILINE)
                text = re.sub(r'^---\s*$', '', text, flags=re.MULTILINE)
                text = re.sub(r'<!--.*?-->', '', text, flags=re.DOTALL)
                text = text.strip()
                if len(text) > 20:
                    sections[sec_no] = text

        # Also try to match by full content search for section numbers not in headers
        for m in re.finditer(r'(?:^|\n)\s*(\d+\.\d+\.[\d.]*)\s+([А-ЯЁЎҚҒҲа-яёўқғҳA-Z])', content):
            sec_no = m.group(1)
            if not sec_no.endswith('.'):
                sec_no += '.'
            if sec_no not in sections:
                # Get text from this position to next section
                start = m.start()
                next_sec = re.search(r'\n\s*\d+\.\d+\.[\d.]*\s+[А-ЯЁЎҚҒҲа-яёўқғҳA-Z]', content[start + 5:])
                end = start + 5 + next_sec.start() if next_sec else min(start + 5000, len(content))
                text = content[start:end].strip()
                if len(text) > 20:
                    sections[sec_no] = text

    return sections


def split_into_sentences(text: str) -> list:
    """Матнни гапларга ажратиш (фарм контекст учун оптимал)."""
    if not text:
        return []

    # Remove table content
    text = re.sub(r'\|[^\n]*\|', '', text)
    # Remove empty lines
    text = re.sub(r'\n{3,}', '\n\n', text)

    # Split into paragraphs first
    paragraphs = text.split('\n\n')
    sentences = []

    for para in paragraphs:
        para = para.strip()
        if not para or len(para) < 10:
            continue
        # Split paragraph into sentences
        # Be careful with abbreviations: мг., мл., г., №, т.д., т.п., бл.
        safe = para
        safe = re.sub(r'(\d+)\.\s', r'\1@@DOT@@', safe)  # 2.1. → protect
        safe = re.sub(r'(мг|мл|мкг|кг|г|нг)\.\s', r'\1@@DOT@@', safe)
        safe = re.sub(r'(т\.д|т\.п|б\.қ|в\.б)\.\s', r'\1@@DOT@@', safe)

        sents = re.split(r'(?<=[.!?])\s+', safe)
        for s in sents:
            s = s.replace('@@DOT@@', '. ').strip()
            if len(s) > 5:
                sentences.append(s)

    return sentences


def build_hatboshilar(toc_entries: list, uz_sections: dict, ru_sections: dict, en_sections: dict):
    """1-ВАЗИФА: Хатбошилар — 3 тилда абзацма-абзац alignment."""
    log.info("=== ХАТБОШИЛАР ===")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Хатбошилар"

    headers = ['№', 'ENGLISH', 'РУССКИЙ', 'ЎЗБЕКСНА', 'МУТАХАССИС', 'МАТН №', 'АМАЛ ТУРИ']
    widths = [6, 55, 55, 55, 20, 10, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, i, h)
        c.font = HF; c.fill = HBG; c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = BD
        ws.column_dimensions[chr(64 + i)].width = w

    row = 2
    total = 0

    for entry in toc_entries:
        mn = entry['matn_no']
        if not mn:
            continue

        # Топиш: UZ ва RU бўлим матнлари
        uz_text = uz_sections.get(mn, '')
        ru_text = ru_sections.get(mn, '')
        en_text = entry['en']  # Мундарижадан

        if not uz_text and not ru_text:
            # Мундарижа сарлавҳасини ўзи ёзиш
            for col, val in [(1, total+1), (2, en_text), (3, entry['ru']), (4, entry['uz']),
                             (5, 'Зайнидинов Акмалходжа'), (6, mn), (7, 'Bulk Save')]:
                c = ws.cell(row, col, val)
                c.font = CF; c.border = BD; c.alignment = WR
            row += 1
            total += 1
            continue

        # Гапларга ажратиш
        uz_sents = split_into_sentences(uz_text)
        ru_sents = split_into_sentences(ru_text)

        # Жуфтлаштириш: мин(uz, ru) гапларгача
        max_sents = max(len(uz_sents), len(ru_sents))
        if max_sents == 0:
            continue

        for si in range(max_sents):
            uz_s = uz_sents[si] if si < len(uz_sents) else ''
            ru_s = ru_sents[si] if si < len(ru_sents) else ''

            for col, val in [
                (1, total + 1),
                (2, en_text if si == 0 else ''),
                (3, ru_s),
                (4, uz_s),
                (5, 'Зайнидинов Акмалходжа'),
                (6, mn),
                (7, 'Bulk Save'),
            ]:
                c = ws.cell(row, col, val)
                c.font = CF; c.border = BD; c.alignment = WR
            row += 1
            total += 1

    out = OUT_DIR / "paragraphs_aligned.xlsx"
    wb.save(str(out))
    log.info(f"Хатбошилар: {total} сатр → {out}")
    return total


def build_izohli(uz_sections: dict, ru_sections: dict):
    """2-ВАЗИФА: Изоҳли луғат — фарм терминлар + таъриф."""
    log.info("=== ИЗОҲЛИ ЛУҒАТ ===")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Изоҳли луғат"

    headers = ['№', 'ENGLISH', 'РУССКИЙ', 'ЎЗБЕКСНА', 'ТАЪРИФ', 'МУТАХАССИС', 'МАТН №']
    widths = [6, 30, 30, 30, 60, 18, 10]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, i, h)
        c.font = HF; c.fill = HBG; c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = BD
        ws.column_dimensions[chr(64 + i)].width = w

    # Терминларни MD матнлардан чиқариш
    # Фарм терминлар: тирноқ ичида, ёки "деб аталади" / "называется" / "defined as" билан
    terms = {}

    # UZ матнлардан
    for sec_no, text in uz_sections.items():
        # "термин" деб аталади pattern
        for m in re.finditer(r'[«"„]([^»""]{3,50})[»""]', text):
            term = m.group(1).strip()
            if len(term) > 3 and term not in terms:
                ctx = text[max(0, m.start()-80):min(len(text), m.end()+120)].replace('\n', ' ')
                terms[term] = {'uz': term, 'context': ctx, 'matn_no': sec_no, 'lang': 'uz'}

    # RU матнлардан
    for sec_no, text in ru_sections.items():
        for m in re.finditer(r'[«"„]([^»""]{3,50})[»""]', text):
            term = m.group(1).strip()
            if len(term) > 3 and term not in terms:
                ctx = text[max(0, m.start()-80):min(len(text), m.end()+120)].replace('\n', ' ')
                terms[term] = {'ru': term, 'context': ctx, 'matn_no': sec_no, 'lang': 'ru'}

    # EP матнлардан
    ep_dir = MD_DIR / "EP_9.0"
    if ep_dir.exists():
        for md_file in sorted(ep_dir.rglob("*.md"))[:200]:
            try:
                content = md_file.read_text(encoding='utf-8', errors='replace')
                for m in re.finditer(r'\b([A-Z][a-z]{2,}(?:\s+[a-z]{2,}){0,2})\s*[:—–]\s*(.{10,100})', content):
                    term = m.group(1)
                    defn = m.group(2).strip()
                    if term not in terms:
                        terms[term] = {'en': term, 'context': defn, 'matn_no': '', 'lang': 'en'}
            except Exception:
                pass

    row = 2
    for i, (key, t) in enumerate(sorted(terms.items()), 1):
        for col, val in [
            (1, i),
            (2, t.get('en', '')),
            (3, t.get('ru', '')),
            (4, t.get('uz', '')),
            (5, t.get('context', '')[:250]),
            (6, 'auto_extract'),
            (7, t.get('matn_no', '')),
        ]:
            c = ws.cell(row, col, val)
            c.font = CF; c.border = BD; c.alignment = WR
        row += 1

    out = OUT_DIR / "annotated_terms.xlsx"
    wb.save(str(out))
    log.info(f"Изоҳли луғат: {len(terms)} термин → {out}")
    return len(terms)


def build_munozarali(uz_sections: dict, ru_sections: dict):
    """3-ВАЗИФА: Мунозарали терминлар."""
    log.info("=== МУНОЗАРАЛИ ===")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мунозарали"

    headers = ['№', 'ENGLISH', 'РУССКИЙ', 'ЎЗБЕКСНА', 'КОНТЕКСТ', 'МУТАХАССИС', 'МАТН №']
    widths = [6, 30, 30, 30, 60, 18, 10]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, i, h)
        c.font = HF; c.fill = HBG; c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = BD
        ws.column_dimensions[chr(64 + i)].width = w

    # Мунозарали: бир хил маъноли лекин турлича ёзиладиган сўзлар
    # ёки ёки/или/or pattern'лар
    disputed = {}

    for sec_no, text in uz_sections.items():
        # "X ёки Y" pattern
        for m in re.finditer(r'(\b\w{3,20})\s+ёки\s+(\w{3,20})\b', text):
            key = f"{m.group(1)}/{m.group(2)}"
            if key not in disputed:
                ctx = text[max(0, m.start()-60):min(len(text), m.end()+80)].replace('\n', ' ')
                disputed[key] = {'uz': key, 'context': ctx, 'matn_no': sec_no}

    for sec_no, text in ru_sections.items():
        for m in re.finditer(r'(\b\w{3,20})\s+или\s+(\w{3,20})\b', text):
            key = f"{m.group(1)}/{m.group(2)}"
            if key not in disputed:
                ctx = text[max(0, m.start()-60):min(len(text), m.end()+80)].replace('\n', ' ')
                disputed[key] = {'ru': key, 'context': ctx, 'matn_no': sec_no}

    row = 2
    for i, (key, t) in enumerate(sorted(disputed.items()), 1):
        for col, val in [
            (1, i), (2, ''), (3, t.get('ru', '')), (4, t.get('uz', '')),
            (5, t.get('context', '')[:250]), (6, 'auto_extract'), (7, t.get('matn_no', '')),
        ]:
            c = ws.cell(row, col, val)
            c.font = CF; c.border = BD; c.alignment = WR
        row += 1

    out = OUT_DIR / "disputed_terms.xlsx"
    wb.save(str(out))
    log.info(f"Мунозарали: {len(disputed)} → {out}")
    return len(disputed)


def build_qisqartmalar(uz_sections: dict, ru_sections: dict):
    """4-ВАЗИФА: Қисқартмалар."""
    log.info("=== ҚИСҚАРТМАЛАР ===")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Қисқартмалар"

    headers = ['№', 'ҚИСҚАРТМА', 'РУССКИЙ', 'ЎЗБЕКСНА', 'ENGLISH FULL NAME', 'МУТАХАССИС', 'МАТН №']
    widths = [6, 15, 40, 40, 40, 18, 10]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, i, h)
        c.font = HF; c.fill = HBG; c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = BD
        ws.column_dimensions[chr(64 + i)].width = w

    abbrs = {}

    # UZ: "ҚИСҚАРТМА — тўлиқ ном" ёки "тўлиқ ном (ҚИСҚАРТМА)" pattern
    for sec_no, text in uz_sections.items():
        # Pattern: (ABBR) after full form
        for m in re.finditer(r'([А-ЯЁЎҚҒҲа-яёўқғҳ\s]{5,60})\s*\(([A-ZА-ЯЁЎҚҒҲ]{2,8})\)', text):
            full = m.group(1).strip()
            short = m.group(2).strip()
            if short not in abbrs and len(full) > 5:
                abbrs[short] = {'short': short, 'uz_full': full, 'matn_no': sec_no}
        # Pattern: ABBR — full form
        for m in re.finditer(r'\b([A-ZА-ЯЁЎҚҒҲ]{2,8})\s*[–—-]\s*([А-ЯЁЎҚҒҲа-яёўқғҳ\s]{5,60})', text):
            short = m.group(1).strip()
            full = m.group(2).strip()
            if short not in abbrs:
                abbrs[short] = {'short': short, 'uz_full': full, 'matn_no': sec_no}

    # RU
    for sec_no, text in ru_sections.items():
        for m in re.finditer(r'([А-Яа-я\s]{5,60})\s*\(([A-ZА-Я]{2,8})\)', text):
            full = m.group(1).strip()
            short = m.group(2).strip()
            if short not in abbrs:
                abbrs[short] = {'short': short, 'ru_full': full, 'matn_no': sec_no}
            elif 'ru_full' not in abbrs[short]:
                abbrs[short]['ru_full'] = full

    # EN (EP)
    ep_dir = MD_DIR / "EP_9.0"
    if ep_dir.exists():
        for md_file in sorted(ep_dir.rglob("*.md"))[:200]:
            try:
                content = md_file.read_text(encoding='utf-8', errors='replace')
                for m in re.finditer(r'([A-Za-z\s]{5,60})\s*\(([A-Z]{2,8})\)', content):
                    full = m.group(1).strip()
                    short = m.group(2).strip()
                    if short in abbrs and 'en_full' not in abbrs[short]:
                        abbrs[short]['en_full'] = full
                    elif short not in abbrs:
                        abbrs[short] = {'short': short, 'en_full': full, 'matn_no': ''}
            except Exception:
                pass

    row = 2
    for i, (key, t) in enumerate(sorted(abbrs.items()), 1):
        for col, val in [
            (1, i),
            (2, t.get('short', '')),
            (3, t.get('ru_full', '')),
            (4, t.get('uz_full', '')),
            (5, t.get('en_full', '')),
            (6, 'auto_extract'),
            (7, t.get('matn_no', '')),
        ]:
            c = ws.cell(row, col, val)
            c.font = CF; c.border = BD; c.alignment = WR
            if col == 2:
                c.font = Font(name='Arial', bold=True, size=10, color='DC2626')
        row += 1

    out = OUT_DIR / "abbreviations.xlsx"
    wb.save(str(out))
    log.info(f"Қисқартмалар: {len(abbrs)} → {out}")
    return len(abbrs)


def main():
    log.info("=== МИССИЯ Д — СИФАТЛИ МАЪЛУМОТЛАР БАЗАСИ ===")

    toc = load_toc()
    log.info(f"Мундарижа: {len(toc)} сатр")

    # MD файллардан бўлимларни юклаш
    log.info("MD файлларни индекслаш...")
    uz = load_all_md("DF_uzb")
    log.info(f"  UZ бўлимлар: {len(uz)}")
    ru = load_all_md("DF_rus")
    log.info(f"  RU бўлимлар: {len(ru)}")
    en = {}  # EN бўлимлар EP'дан кейинроқ

    n1 = build_hatboshilar(toc, uz, ru, en)
    n2 = build_izohli(uz, ru)
    n3 = build_munozarali(uz, ru)
    n4 = build_qisqartmalar(uz, ru)

    log.info(f"\n=== НАТИЖА ===")
    log.info(f"1. Хатбошилар:   {n1} сатр")
    log.info(f"2. Изоҳли луғат: {n2} термин")
    log.info(f"3. Мунозарали:   {n3} термин")
    log.info(f"4. Қисқартмалар: {n4} та")


if __name__ == "__main__":
    main()
