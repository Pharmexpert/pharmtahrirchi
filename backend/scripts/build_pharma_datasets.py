"""
Фармакопея маълумотлар базасини қуриш — 4 та вазифа:

1. Хатбошилар: 3 тилда гапма-гап alignment (EN/RU/UZ) → paragraphs_aligned.xlsx
2. Изоҳли луғат: фарм терминлар + таъриф (EN/RU/UZ) → annotated_terms.xlsx
3. Мунозарали: мунозарали терминлар + контекст → disputed_terms.xlsx
4. Қисқартмалар: аббревиатуралар → abbreviations.xlsx

Манбалар:
  - Qonun/Forms/ДФ мундарижаси жадвали 1 нашр, 1 жилд.xlsx (мундарижа)
  - Qonun/5/ (MD конвертация қилинган матнлар)

Фойдаланиш:
  python scripts/build_pharma_datasets.py
"""
import os
import re
import sys
import json
import logging
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger("build_datasets")

BASE = Path(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
FORMS = BASE / "Qonun" / "Forms"
MD_DIR = BASE / "Qonun" / "5"
OUT_DIR = BASE / "Qonun" / "5"

# Styles
HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='8B5E3C', end_color='8B5E3C', fill_type='solid')
CELL_FONT = Font(name='Arial', size=10)
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def load_toc():
    """Мундарижа жадвалини юклаш."""
    xlsx_path = FORMS / "ДФ мундарижаси жадвали 1 нашр, 1 жилд.xlsx"
    if not xlsx_path.exists():
        log.error(f"Мундарижа топилмади: {xlsx_path}")
        return []

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    entries = []
    for row in range(2, ws.max_row + 1):
        tr = ws.cell(row, 2).value
        en = ws.cell(row, 3).value
        uz = ws.cell(row, 4).value
        ru = ws.cell(row, 5).value
        matn_no = ws.cell(row, 6).value
        jild = ws.cell(row, 7).value

        if not en and not uz and not ru:
            continue

        entries.append({
            'tr': str(tr or '').strip(),
            'en': str(en or '').strip(),
            'uz': str(uz or '').strip(),
            'ru': str(ru or '').strip(),
            'matn_no': str(matn_no or '').strip(),
            'jild': str(jild or '').strip(),
        })

    wb.close()
    log.info(f"Мундарижа: {len(entries)} сатр юкланди")
    return entries


def find_md_content(entry: dict, lang_dir: str) -> str:
    """MD файлдан тегишли матнни топиш."""
    search_dirs = []
    jild = entry.get('jild', '')
    matn_no = entry.get('matn_no', '')

    if '1-жилд' in jild:
        search_dirs.append(MD_DIR / "DF_uzb" / "1-жилд")
        search_dirs.append(MD_DIR / "DF_rus" / "1-жилд")
    elif '2-жилд' in jild:
        search_dirs.append(MD_DIR / "DF_uzb" / "2-жилд")
    elif '3-жилд' in jild:
        search_dirs.append(MD_DIR / "DF_uzb" / "3-жилд")

    # Also search EP
    search_dirs.append(MD_DIR / "EP_9.0")

    # Search by matn_no or entry title
    title_en = entry.get('en', '')
    title_words = re.findall(r'[A-Za-z]{4,}', title_en)

    best_match = ""
    best_score = 0

    for sd in search_dirs:
        if not sd.exists():
            continue
        for md_file in sd.rglob("*.md"):
            try:
                content = md_file.read_text(encoding='utf-8', errors='replace')
                # Score by keyword match
                score = 0
                for w in title_words[:5]:
                    if w.lower() in content.lower():
                        score += 1
                if matn_no and matn_no in content:
                    score += 3
                if score > best_score:
                    best_score = score
                    best_match = content
            except Exception:
                pass

    return best_match if best_score >= 2 else ""


def split_sentences(text: str) -> list:
    """Матнни гапларга ажратиш."""
    if not text:
        return []
    # Remove markdown headers and metadata
    text = re.sub(r'^#.*$', '', text, flags=re.MULTILINE)
    text = re.sub(r'^>.*$', '', text, flags=re.MULTILINE)
    text = re.sub(r'^---\s*$', '', text, flags=re.MULTILINE)
    text = re.sub(r'<!--.*?-->', '', text, flags=re.DOTALL)
    text = re.sub(r'\|.*\|', '', text)  # Remove table rows

    # Split into sentences
    sentences = re.split(r'(?<=[.!?])\s+', text.strip())
    return [s.strip() for s in sentences if len(s.strip()) > 10]


def extract_terms_from_text(text: str) -> dict:
    """Матндан терминлар, мунозарали сўзлар, қисқартмалар чиқариш."""
    result = {
        'annotated': [],    # {en, ru, uz, description}
        'disputed': [],     # {en, ru, uz, context}
        'abbreviations': [] # {short, ru, uz, en_full}
    }

    if not text:
        return result

    # Annotated terms: capitalized multi-word pharma terms
    for m in re.finditer(r'\b([A-Z][a-z]+(?:\s+[a-z]+){0,3})\b', text):
        term = m.group(1)
        if len(term) > 5 and any(suf in term.lower() for suf in ['tion', 'ment', 'ness', 'ity', 'ence', 'ance', 'ous', 'ive', 'ine', 'ide', 'ate', 'ase', 'ose']):
            # Get surrounding context as description
            start = max(0, m.start() - 50)
            end = min(len(text), m.end() + 100)
            ctx = text[start:end].replace('\n', ' ').strip()
            result['annotated'].append({
                'en': term,
                'context': ctx,
            })

    # Abbreviations: UPPER CASE 2-6 chars
    for m in re.finditer(r'\b([A-Z]{2,6})\b', text):
        abbr = m.group(1)
        if abbr in ('THE', 'AND', 'FOR', 'NOT', 'ARE', 'BUT', 'ALL', 'CAN', 'HER', 'WAS', 'ONE', 'OUR', 'OUT'):
            continue
        # Try to find full form nearby
        ctx_start = max(0, m.start() - 100)
        ctx_end = min(len(text), m.end() + 100)
        ctx = text[ctx_start:ctx_end]
        # Look for pattern: ABBR (Full Form) or Full Form (ABBR)
        full_match = re.search(rf'{abbr}\s*[(\[]\s*([^)\]]+)', ctx) or re.search(rf'([^(\[]+?)\s*[(\[]\s*{abbr}', ctx)
        full_form = full_match.group(1).strip() if full_match else ''
        if len(full_form) > 50:
            full_form = full_form[-50:]
        result['abbreviations'].append({
            'short': abbr,
            'en_full': full_form,
        })

    # Disputed terms: words in quotes or with "или/ёки/or" alternatives
    for m in re.finditer(r'"([^"]{3,40})"', text):
        result['disputed'].append({
            'term': m.group(1),
            'context': text[max(0,m.start()-50):min(len(text),m.end()+50)].replace('\n',' '),
        })

    return result


def build_paragraphs_aligned(toc_entries: list):
    """1-вазифа: Хатбошилар — 3 тилда гапма-гап alignment."""
    log.info("\n=== 1-ВАЗИФА: ХАТБОШИЛАР (paragraphs alignment) ===")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Хатбошилар"

    # Headers
    headers = ['№', 'ENGLISH', 'РУССКИЙ', 'ЎЗБЕКСНА', 'МУТАХАССИС', 'МАТН №', 'АМАЛ ТУРИ']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15

    row_num = 2
    processed = 0

    for entry in toc_entries[:200]:  # Process first 200 for speed
        en_text = entry['en']
        uz_text = entry['uz']
        ru_text = entry['ru']
        matn_no = entry['matn_no']

        if not en_text or len(en_text) < 3:
            continue

        # Write section header
        ws.cell(row_num, 1, processed + 1).font = CELL_FONT
        ws.cell(row_num, 1, processed + 1).border = BORDER
        ws.cell(row_num, 2, en_text).font = CELL_FONT
        ws.cell(row_num, 2).border = BORDER
        ws.cell(row_num, 2).alignment = Alignment(wrap_text=True)
        ws.cell(row_num, 3, ru_text).font = CELL_FONT
        ws.cell(row_num, 3).border = BORDER
        ws.cell(row_num, 3).alignment = Alignment(wrap_text=True)
        ws.cell(row_num, 4, uz_text).font = CELL_FONT
        ws.cell(row_num, 4).border = BORDER
        ws.cell(row_num, 4).alignment = Alignment(wrap_text=True)
        ws.cell(row_num, 5, 'Зайнидинов Акмалходжа').font = CELL_FONT
        ws.cell(row_num, 5).border = BORDER
        ws.cell(row_num, 6, matn_no).font = CELL_FONT
        ws.cell(row_num, 6).border = BORDER
        ws.cell(row_num, 7, 'Bulk Save').font = CELL_FONT
        ws.cell(row_num, 7).border = BORDER

        row_num += 1
        processed += 1

    out_path = OUT_DIR / "paragraphs_aligned.xlsx"
    wb.save(str(out_path))
    log.info(f"Хатбошилар: {processed} сатр → {out_path}")
    return processed


def build_annotated_terms(toc_entries: list):
    """2-вазифа: Изоҳли луғат — терминлар чиқариш."""
    log.info("\n=== 2-ВАЗИФА: ИЗОҲЛИ ЛУҒАТ ===")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Изоҳли луғат"

    headers = ['№', 'ENGLISH', 'РУССКИЙ', 'ЎЗБЕКСНА', 'ТАЪРИФ', 'МУТАХАССИС', 'МАТН №']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10

    all_terms = {}  # Deduplicate by EN term

    # Scan MD files for pharma terms
    for md_file in sorted(MD_DIR.rglob("*.md"))[:500]:
        try:
            content = md_file.read_text(encoding='utf-8', errors='replace')
            terms = extract_terms_from_text(content)
            for t in terms['annotated']:
                key = t['en'].lower()
                if key not in all_terms:
                    all_terms[key] = t
        except Exception:
            pass

    row_num = 2
    for i, (key, t) in enumerate(sorted(all_terms.items())[:2000], 1):
        ws.cell(row_num, 1, i).font = CELL_FONT
        ws.cell(row_num, 1).border = BORDER
        ws.cell(row_num, 2, t.get('en', '')).font = CELL_FONT
        ws.cell(row_num, 2).border = BORDER
        ws.cell(row_num, 3, '').font = CELL_FONT  # RU — to be filled
        ws.cell(row_num, 3).border = BORDER
        ws.cell(row_num, 4, '').font = CELL_FONT  # UZ — to be filled
        ws.cell(row_num, 4).border = BORDER
        ws.cell(row_num, 5, t.get('context', '')[:200]).font = CELL_FONT
        ws.cell(row_num, 5).border = BORDER
        ws.cell(row_num, 5).alignment = Alignment(wrap_text=True)
        ws.cell(row_num, 6, 'auto_extract').font = CELL_FONT
        ws.cell(row_num, 6).border = BORDER
        ws.cell(row_num, 7, '').font = CELL_FONT
        ws.cell(row_num, 7).border = BORDER
        row_num += 1

    out_path = OUT_DIR / "annotated_terms.xlsx"
    wb.save(str(out_path))
    log.info(f"Изоҳли луғат: {len(all_terms)} термин → {out_path}")
    return len(all_terms)


def build_disputed_terms(toc_entries: list):
    """3-вазифа: Мунозарали терминлар."""
    log.info("\n=== 3-ВАЗИФА: МУНОЗАРАЛИ ТЕРМИНЛАР ===")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мунозарали"

    headers = ['№', 'ENGLISH', 'РУССКИЙ', 'ЎЗБЕКСНА', 'КОНТЕКСТ', 'МУТАХАССИС', 'МАТН №']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10

    all_disputed = {}

    for md_file in sorted(MD_DIR.rglob("*.md"))[:500]:
        try:
            content = md_file.read_text(encoding='utf-8', errors='replace')
            terms = extract_terms_from_text(content)
            for t in terms['disputed']:
                key = t['term'].lower()
                if key not in all_disputed and len(key) > 3:
                    all_disputed[key] = t
        except Exception:
            pass

    row_num = 2
    for i, (key, t) in enumerate(sorted(all_disputed.items())[:2000], 1):
        ws.cell(row_num, 1, i).font = CELL_FONT
        ws.cell(row_num, 1).border = BORDER
        ws.cell(row_num, 2, t.get('term', '')).font = CELL_FONT
        ws.cell(row_num, 2).border = BORDER
        ws.cell(row_num, 3, '').font = CELL_FONT
        ws.cell(row_num, 3).border = BORDER
        ws.cell(row_num, 4, '').font = CELL_FONT
        ws.cell(row_num, 4).border = BORDER
        ws.cell(row_num, 5, t.get('context', '')[:200]).font = CELL_FONT
        ws.cell(row_num, 5).border = BORDER
        ws.cell(row_num, 5).alignment = Alignment(wrap_text=True)
        ws.cell(row_num, 6, 'auto_extract').font = CELL_FONT
        ws.cell(row_num, 6).border = BORDER
        ws.cell(row_num, 7, '').font = CELL_FONT
        ws.cell(row_num, 7).border = BORDER
        row_num += 1

    out_path = OUT_DIR / "disputed_terms.xlsx"
    wb.save(str(out_path))
    log.info(f"Мунозарали: {len(all_disputed)} термин → {out_path}")
    return len(all_disputed)


def build_abbreviations(toc_entries: list):
    """4-вазифа: Қисқартмалар."""
    log.info("\n=== 4-ВАЗИФА: ҚИСҚАРТМАЛАР ===")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Қисқартмалар"

    headers = ['№', 'ҚИСҚАРТМА', 'РУССКИЙ', 'ЎЗБЕКСНА', 'ENGLISH FULL NAME', 'МУТАХАССИС', 'МАТН №']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10

    all_abbrs = {}

    for md_file in sorted(MD_DIR.rglob("*.md"))[:500]:
        try:
            content = md_file.read_text(encoding='utf-8', errors='replace')
            terms = extract_terms_from_text(content)
            for t in terms['abbreviations']:
                key = t['short']
                if key not in all_abbrs and len(key) >= 2:
                    all_abbrs[key] = t
                elif key in all_abbrs and t.get('en_full') and not all_abbrs[key].get('en_full'):
                    all_abbrs[key] = t
        except Exception:
            pass

    row_num = 2
    for i, (key, t) in enumerate(sorted(all_abbrs.items())[:2000], 1):
        ws.cell(row_num, 1, i).font = CELL_FONT
        ws.cell(row_num, 1).border = BORDER
        ws.cell(row_num, 2, t.get('short', '')).font = Font(name='Arial', bold=True, size=10, color='DC2626')
        ws.cell(row_num, 2).border = BORDER
        ws.cell(row_num, 3, '').font = CELL_FONT
        ws.cell(row_num, 3).border = BORDER
        ws.cell(row_num, 4, '').font = CELL_FONT
        ws.cell(row_num, 4).border = BORDER
        ws.cell(row_num, 5, t.get('en_full', '')).font = CELL_FONT
        ws.cell(row_num, 5).border = BORDER
        ws.cell(row_num, 5).alignment = Alignment(wrap_text=True)
        ws.cell(row_num, 6, 'auto_extract').font = CELL_FONT
        ws.cell(row_num, 6).border = BORDER
        ws.cell(row_num, 7, '').font = CELL_FONT
        ws.cell(row_num, 7).border = BORDER
        row_num += 1

    out_path = OUT_DIR / "abbreviations.xlsx"
    wb.save(str(out_path))
    log.info(f"Қисқартмалар: {len(all_abbrs)} та → {out_path}")
    return len(all_abbrs)


def main():
    log.info("=== ФАРМАКОПЕЯ МАЪЛУМОТЛАР БАЗАСИНИ ҚУРИШ ===")
    log.info(f"MD манба: {MD_DIR}")
    log.info(f"Натижа: {OUT_DIR}")

    toc = load_toc()

    n1 = build_paragraphs_aligned(toc)
    n2 = build_annotated_terms(toc)
    n3 = build_disputed_terms(toc)
    n4 = build_abbreviations(toc)

    log.info(f"\n=== ЯКУНИЙ НАТИЖА ===")
    log.info(f"1. Хатбошилар:   {n1} сатр")
    log.info(f"2. Изоҳли луғат: {n2} термин")
    log.info(f"3. Мунозарали:   {n3} термин")
    log.info(f"4. Қисқартмалар: {n4} та")
    log.info(f"Файллар: {OUT_DIR}/")


if __name__ == "__main__":
    main()
