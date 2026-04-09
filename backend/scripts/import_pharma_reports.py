"""
Import 3 pharmacopoeia report XLSX files into the UI-facing tables:
  1. glossary_report.xlsx
       Изоҳли сўзлар → annotated_words (uz=term, description_uz=definition)
       Қисқартмалар + Стандарт рўйхат → abbreviations (short_form, long_uz)
  2. jildlar_tahlil_hisoboti.xlsx → sayqallash_rules (per volume)
  3. linguistic_analysis_report.xlsx → sayqallash_rules (summary)

All inserts use INSERT OR IGNORE so it is safe to re-run.
"""
import os
import sqlite3
import logging

log = logging.getLogger("import_pharma_reports")

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "source_data", "pharmacopoeia")
DB_PATH = os.getenv("DB_PATH", os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "pharma_editor.db"))


def _ensure_schemas(cur):
    # All these tables are also created in db.init_db on startup; this just makes
    # the script standalone-runnable on a fresh DB.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS annotated_words (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            en TEXT, ru TEXT, uz TEXT,
            description_en TEXT, description_ru TEXT, description_uz TEXT,
            source_lang TEXT DEFAULT 'Uzbek',
            user_id TEXT, text_id TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            modified_by_id TEXT, modified_at TIMESTAMP,
            status TEXT DEFAULT 'active'
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS abbreviations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            short_form TEXT UNIQUE NOT NULL,
            long_en TEXT, long_ru TEXT, long_uz TEXT,
            source_lang TEXT DEFAULT 'Uzbek',
            user_id TEXT, text_id TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            modified_by_id TEXT, modified_at TIMESTAMP,
            status TEXT DEFAULT 'active'
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sayqallash_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            wrong_form TEXT NOT NULL,
            correct_form TEXT NOT NULL,
            error_type TEXT,
            frequency INTEGER DEFAULT 1,
            lang TEXT DEFAULT 'uz',
            source TEXT,
            context TEXT,
            UNIQUE(wrong_form, correct_form)
        )
    """)
    # Migration: ensure annotated_words has 'source' column for traceability
    try:
        cur.execute("ALTER TABLE annotated_words ADD COLUMN source TEXT DEFAULT ''")
    except Exception:
        pass


def _import_glossary(cur) -> dict:
    path = os.path.join(DATA_DIR, "glossary_report.xlsx")
    if not os.path.exists(path):
        return {"definitions": 0, "abbreviations": 0, "error": "file_not_found"}
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl_missing"}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    defs_ins = 0
    abbr_ins = 0

    # Sheet 1: Изоҳли сўзлар → annotated_words
    if "Изоҳли сўзлар" in wb.sheetnames:
        ws = wb["Изоҳли сўзлар"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or len(row) < 3:
                continue
            _, term, definition, *rest = list(row) + [None] * 5
            if not term or not definition:
                continue
            term_s = str(term).strip()
            def_s = str(definition).strip()
            try:
                # Dedupe by (uz term, source)
                cur.execute(
                    "SELECT 1 FROM annotated_words WHERE uz = ? AND source = 'ДФ глоссарий' LIMIT 1",
                    (term_s,),
                )
                if cur.fetchone():
                    continue
                cur.execute(
                    """INSERT INTO annotated_words
                       (uz, description_uz, source_lang, source, status)
                       VALUES (?, ?, 'Uzbek', 'ДФ глоссарий', 'active')""",
                    (term_s, def_s),
                )
                defs_ins += 1
            except Exception as e:
                log.debug("def row skip: %s", e)

    # Sheets 2 & 3 → abbreviations table
    for sheet in ("Қисқартмалар", "Стандарт рўйхат"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or len(row) < 3:
                continue
            _, abbr, full, *_ = list(row) + [None] * 3
            if not abbr or not full:
                continue
            try:
                cur.execute(
                    """INSERT OR IGNORE INTO abbreviations
                       (short_form, long_uz, source_lang, status)
                       VALUES (?, ?, 'Uzbek', 'active')""",
                    (str(abbr).strip(), str(full).strip()),
                )
                abbr_ins += cur.rowcount
            except Exception as e:
                log.debug("abbr row skip: %s", e)
    wb.close()
    return {"definitions": defs_ins, "abbreviations": abbr_ins}


def _import_error_xlsx(cur, path: str, source_label: str) -> int:
    if not os.path.exists(path):
        return 0
    try:
        import openpyxl
    except ImportError:
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    inserted = 0
    for sheet in wb.sheetnames:
        if sheet.upper().startswith("ЖАМИ"):
            continue
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or not row or len(row) < 4:
                continue
            _, wrong, correct, etype, *rest = list(row) + [None] * 5
            freq = rest[0] if rest else 1
            if not wrong or not correct:
                continue
            try:
                cur.execute(
                    """INSERT OR IGNORE INTO sayqallash_rules
                       (wrong_form, correct_form, error_type, frequency, lang, source, context)
                       VALUES (?,?,?,?,'uz',?,?)""",
                    (
                        str(wrong).strip(),
                        str(correct).strip(),
                        str(etype or "Stilistik")[:80],
                        int(freq) if isinstance(freq, (int, float)) else 1,
                        source_label,
                        f"{source_label} / {sheet}",
                    ),
                )
                inserted += cur.rowcount
            except Exception as e:
                log.debug("err row skip: %s", e)
    wb.close()
    return inserted


def main() -> dict:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    _ensure_schemas(cur)

    glossary = _import_glossary(cur)
    jildlar = _import_error_xlsx(cur, os.path.join(DATA_DIR, "jildlar_tahlil_hisoboti.xlsx"), "ДФ жилдлар таҳлили")
    linguistic = _import_error_xlsx(cur, os.path.join(DATA_DIR, "linguistic_analysis_report.xlsx"), "ДФ лингвистик ҳисобот")

    conn.commit()
    conn.close()

    result = {
        "annotated_words_added": glossary.get("definitions", 0),
        "abbreviations_added": glossary.get("abbreviations", 0),
        "sayqallash_from_jildlar": jildlar,
        "sayqallash_from_linguistic": linguistic,
        "total_sayqallash": jildlar + linguistic,
    }
    log.info("import_pharma_reports result: %s", result)
    return result


if __name__ == "__main__":
    print(main())
